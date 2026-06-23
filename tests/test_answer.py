from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from draftcode.answer import write_answer_card
from draftcode.io import load_draft_order, load_mock_signals, load_prospects, load_team_needs
from draftcode.simulate import MonteCarloDraftTwin, SimulationConfig, TwinReport

PROCESSED_DIR = Path("data/processed")
TEMPLATE = Path("data/raw/official/answer_card_template.xlsx")


def _run_report(seed: int = 42, draws: int = 80) -> TwinReport:
    return MonteCarloDraftTwin(
        prospects=load_prospects(PROCESSED_DIR),
        draft_order=load_draft_order(PROCESSED_DIR),
        team_needs=load_team_needs(PROCESSED_DIR),
        mock_signals=load_mock_signals(PROCESSED_DIR),
        config=SimulationConfig(draws=draws, seed=seed),
    ).run()


def test_answer_card_written_to_template(tmp_path: Path) -> None:
    report = _run_report()
    out = tmp_path / "answer_card.xlsx"

    write_answer_card(template=TEMPLATE, out=out, report=report, team_id="Team01")

    workbook = load_workbook(out, data_only=True)
    assert _value_right_of(workbook["①队伍信息"], "队伍ID") == "Team01"

    pool_names = {prospect.name for prospect in load_prospects(PROCESSED_DIR)}
    pick_sheet = workbook["②选秀名单预测"]
    pick_header, pick_columns = _find_columns(pick_sheet, ["顺位", "预测球员"])
    filled_names = []
    for pick in range(1, 31):
        row = _find_pick_row(pick_sheet, pick_header, pick_columns["顺位"], pick)
        value = _cell_text(pick_sheet.cell(row=row, column=pick_columns["预测球员"]).value)
        filled_names.append(value)

    assert len(filled_names) == 30
    assert all(filled_names)
    assert set(filled_names) <= pool_names
    assert len(set(filled_names)) == 30

    milestone_answers = _read_milestone_answers(workbook["③里程碑预测"])
    assert set(milestone_answers) == {f"Q{index}" for index in range(1, 8)}
    assert all(milestone_answers.values())
    assert 0 <= int(milestone_answers["Q3"]) <= 30
    assert milestone_answers["Q6"] in {
        prospect.school for prospect in load_prospects(PROCESSED_DIR)
    }


def test_answer_card_deterministic_for_same_seed(tmp_path: Path) -> None:
    first_report = _run_report(seed=42)
    second_report = _run_report(seed=42)
    first_out = tmp_path / "first.xlsx"
    second_out = tmp_path / "second.xlsx"

    write_answer_card(template=TEMPLATE, out=first_out, report=first_report)
    write_answer_card(template=TEMPLATE, out=second_out, report=second_report)

    first_book = load_workbook(first_out, data_only=True)
    second_book = load_workbook(second_out, data_only=True)

    assert _read_pick_answers(first_book["②选秀名单预测"]) == _read_pick_answers(
        second_book["②选秀名单预测"]
    )
    assert _read_milestone_answers(first_book["③里程碑预测"]) == _read_milestone_answers(
        second_book["③里程碑预测"]
    )


def _read_pick_answers(sheet: Worksheet) -> list[str]:
    header_row, columns = _find_columns(sheet, ["顺位", "预测球员"])
    answers: list[str] = []
    for pick in range(1, 31):
        row = _find_pick_row(sheet, header_row, columns["顺位"], pick)
        answers.append(_cell_text(sheet.cell(row=row, column=columns["预测球员"]).value))
    return answers


def _read_milestone_answers(sheet: Worksheet) -> dict[str, str]:
    header_row, columns = _find_columns(sheet, ["题号", "你的答案"])
    answers: dict[str, str] = {}
    for question_id in [f"Q{index}" for index in range(1, 8)]:
        row = _find_text_row(sheet, header_row, columns["题号"], question_id)
        answers[question_id] = _cell_text(sheet.cell(row=row, column=columns["你的答案"]).value)
    return answers


def _find_columns(sheet: Worksheet, labels: list[str]) -> tuple[int, dict[str, int]]:
    for row in sheet.iter_rows():
        columns: dict[str, int] = {}
        for label in labels:
            for cell in row:
                if label in _cell_text(cell.value):
                    columns[label] = cell.column
                    break
        if len(columns) == len(labels):
            return row[0].row, columns
    raise AssertionError(f"Could not locate labels in {sheet.title}")


def _find_pick_row(sheet: Worksheet, header_row: int, column: int, pick: int) -> int:
    for row_index in range(header_row + 1, sheet.max_row + 1):
        if _pick_number(sheet.cell(row=row_index, column=column).value) == pick:
            return row_index
    raise AssertionError(f"Could not locate pick {pick}")


def _find_text_row(sheet: Worksheet, header_row: int, column: int, value: str) -> int:
    for row_index in range(header_row + 1, sheet.max_row + 1):
        if _cell_text(sheet.cell(row=row_index, column=column).value) == value:
            return row_index
    raise AssertionError(f"Could not locate row {value}")


def _value_right_of(sheet: Worksheet, label: str) -> str:
    for row in sheet.iter_rows():
        for cell in row:
            if _cell_text(cell.value) == label:
                return _cell_text(sheet.cell(row=cell.row, column=cell.column + 1).value)
    raise AssertionError(f"Could not locate {label}")


def _pick_number(value: object) -> int | None:
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = _cell_text(value)
    if not text:
        return None
    try:
        parsed = float(text)
    except ValueError:
        return None
    if parsed.is_integer():
        return int(parsed)
    return None


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()
