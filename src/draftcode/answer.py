from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from draftcode.simulate import TwinReport

PICKS_SHEET = "②选秀名单预测"
MILESTONES_SHEET = "③里程碑预测"
TEAM_SHEET = "①队伍信息"
POOL_SHEET = "球员池_勿修改"


def write_answer_card(
    template: Path,
    out: Path,
    report: TwinReport,
    team_id: str | None = None,
) -> None:
    """Fill the official answer-card template from a Draft Twin report."""
    workbook = load_workbook(template)

    _write_pick_predictions(workbook[PICKS_SHEET], report)
    _write_milestone_answers(workbook[MILESTONES_SHEET], report)
    if team_id:
        _write_team_id(workbook[TEAM_SHEET], team_id)

    out.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(out)


def _write_pick_predictions(sheet: Worksheet, report: TwinReport) -> None:
    header_row, columns = _find_columns(sheet, ["顺位", "持签球队", "预测球员"])
    pick_column = columns["顺位"]
    team_column = columns["持签球队"]
    answer_column = columns["预测球员"]

    for pick in report.assigned_picks:
        row = _find_pick_row(sheet, header_row, pick_column, pick.pick)
        # Holding team comes from our draft order (official board + officially
        # confirmed trades, e.g. the Giannis deal moving pick 13 to Milwaukee),
        # overriding the template's pre-trade pre-fill.
        sheet.cell(row=row, column=team_column).value = pick.team
        sheet.cell(row=row, column=answer_column).value = pick.prospect_name


def _write_milestone_answers(sheet: Worksheet, report: TwinReport) -> None:
    header_row, columns = _find_columns(sheet, ["题号", "你的答案"])
    id_column = columns["题号"]
    answer_column = columns["你的答案"]
    milestones = {milestone.id: milestone for milestone in report.milestones}

    for question_id in [f"Q{index}" for index in range(1, 8)]:
        if question_id not in milestones:
            raise ValueError(f"Missing milestone answer: {question_id}")
        row = _find_text_row(sheet, header_row, id_column, question_id)
        milestone = milestones[question_id]
        # Submit the board-consistent answer so the card never contradicts the
        # predicted 30-pick board; fall back to the distribution answer if unset.
        sheet.cell(row=row, column=answer_column).value = (
            milestone.board_answer_display or milestone.answer_display
        )


def _write_team_id(sheet: Worksheet, team_id: str) -> None:
    for row in sheet.iter_rows():
        for cell in row:
            if _cell_text(cell.value) == "队伍ID":
                sheet.cell(row=cell.row, column=cell.column + 1).value = team_id
                return
    raise ValueError(f"Could not locate 队伍ID row in sheet {sheet.title!r}")


def _find_columns(sheet: Worksheet, labels: list[str]) -> tuple[int, dict[str, int]]:
    for row in sheet.iter_rows():
        columns: dict[str, int] = {}
        for label in labels:
            for cell in row:
                if label in _cell_text(cell.value):
                    columns[label] = cell.column
                    break
        if len(columns) == len(labels):
            return row[0].row, columns
    label_text = ", ".join(labels)
    raise ValueError(f"Could not locate header columns {label_text} in sheet {sheet.title!r}")


def _find_pick_row(sheet: Worksheet, header_row: int, column: int, pick: int) -> int:
    for row_index in range(header_row + 1, sheet.max_row + 1):
        value = sheet.cell(row=row_index, column=column).value
        if _pick_number(value) == pick:
            return row_index
    raise ValueError(f"Could not locate pick {pick} in sheet {sheet.title!r}")


def _find_text_row(sheet: Worksheet, header_row: int, column: int, value: str) -> int:
    for row_index in range(header_row + 1, sheet.max_row + 1):
        if _cell_text(sheet.cell(row=row_index, column=column).value) == value:
            return row_index
    raise ValueError(f"Could not locate row value {value!r} in sheet {sheet.title!r}")


def _pick_number(value: object) -> int | None:
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = _cell_text(value)
    if not text:
        return None
    try:
        parsed = float(text)
    except ValueError:
        return None
    if parsed.is_integer():
        return int(parsed)
    return None


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()
