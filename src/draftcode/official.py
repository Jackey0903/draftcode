from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from draftcode import divergence
from draftcode.dossier import TeamDossier, load_team_dossiers

ZERO_WIDTH_NON_JOINER = "\u200c"
FEET_INCHES_RE = re.compile(r"(\d+)'\s*([\d.]+)''?")
MISSING_TEXT = {"", "-", "-%", "none", "nan", "n/a", "#n/a"}
DEFAULT_DOSSIER_PATH = Path("data/dossiers/team_dossiers.json")

PROSPECT_COLUMNS = [
    "prospect_id",
    "name",
    "primary_position",
    "archetype",
    "consensus_rank",
    "age",
    "height_in",
    "wingspan_in",
    "usage_rate",
    "true_shooting_pct",
    "assist_rate",
    "rebound_rate",
    "stock_rate",
    "notes",
    "barefoot_height_in",
    "hand_length_in",
    "hand_width_in",
    "standing_reach_in",
    "weight_lb",
    "max_vertical_in",
    "standing_vertical_in",
    "school",
    "country",
    "is_international",
    "is_center",
    "talent_composite",
    "espn_rank",
    "model_pick_low",
    "board_source",
    "talent_rank",
    "market_rank",
    "talent_signal",
    "market_signal",
    "divergence_gap",
    "divergence_type",
    "divergence_reason",
    "divergence_llm_verdict",
    "divergence_llm_market_weight",
    "divergence_llm_confidence",
    "divergence_llm_reasoning",
    "fused_score",
]

DRAFT_ORDER_COLUMNS = ["pick", "team", "abbreviation"]
TEAM_NEEDS_COLUMNS = ["abbreviation", "position", "weight"]
MOCK_SIGNALS_COLUMNS = ["abbreviation", "prospect_id", "signal_strength", "source"]

POSITION_TO_PRIMARY = {
    "后卫": "G",
    "PG": "G",
    "SG": "G",
    "前锋": "W",
    "SF": "W",
    "PF": "W",
    "中锋": "B",
    "C": "B",
}

ARCHETYPES = {
    "G": "guard",
    "W": "wing",
    "B": "big",
}

POSITION_DEFAULTS = {
    "G": {
        "height_in": 74.0,
        "wingspan_in": 78.0,
        "usage_rate": 24.0,
        "true_shooting_pct": 0.55,
        "assist_rate": 22.0,
        "rebound_rate": 9.0,
        "stock_rate": 2.5,
    },
    "W": {
        "height_in": 79.0,
        "wingspan_in": 83.0,
        "usage_rate": 22.0,
        "true_shooting_pct": 0.56,
        "assist_rate": 12.0,
        "rebound_rate": 12.0,
        "stock_rate": 3.0,
    },
    "B": {
        "height_in": 83.0,
        "wingspan_in": 88.0,
        "usage_rate": 20.0,
        "true_shooting_pct": 0.58,
        "assist_rate": 8.0,
        "rebound_rate": 20.0,
        "stock_rate": 4.0,
    },
}

COUNTRY_NAMES = {
    "墨西哥",
    "西班牙",
    "法国",
    "德国",
    "意大利",
    "塞尔维亚",
    "澳大利亚",
    "俄罗斯",
    "塞内加尔",
    "伊朗",
    "克罗地亚",
    "芬兰",
    "英国",
    "爱沙尼亚",
    "以色列",
    "尼日利亚",
}

TEAM_ABBREVIATIONS = {
    "Atlanta Hawks": "ATL",
    "Boston Celtics": "BOS",
    "Brooklyn Nets": "BKN",
    "Charlotte Hornets": "CHA",
    "Chicago Bulls": "CHI",
    "Cleveland Cavaliers": "CLE",
    "Dallas Mavericks": "DAL",
    "Denver Nuggets": "DEN",
    "Detroit Pistons": "DET",
    "Golden State Warriors": "GSW",
    "Los Angeles Clippers": "LAC",
    "Los Angeles Lakers": "LAL",
    "Memphis Grizzlies": "MEM",
    "Miami Heat": "MIA",
    "Milwaukee Bucks": "MIL",
    "Minnesota Timberwolves": "MIN",
    "New York Knicks": "NYK",
    "Oklahoma City Thunder": "OKC",
    "Philadelphia 76ers": "PHI",
    "Sacramento Kings": "SAC",
    "San Antonio Spurs": "SAS",
    "Toronto Raptors": "TOR",
    "Utah Jazz": "UTA",
    "Washington Wizards": "WAS",
}

SHOOTING_PAIRS = [
    ("运球投篮命中", "运球投篮出手"),
    ("定点投篮命中", "定点投篮出手"),
    ("三分定点命中", "三分定点出手"),
    ("中距离定点命中", "中距离定点出手"),
    ("三分侧翼命中", "三分侧翼出手"),
    ("中距离侧翼命中", "中距离侧翼出手"),
    ("罚球命中", "罚球出手"),
]


@dataclass(frozen=True)
class PoolProspect:
    pool_index: int
    prospect_id: str
    name: str
    position: str
    school: str
    raw_country: str


@dataclass(frozen=True)
class HandbookMapping:
    english_name: str
    pool_index: int
    chinese_name: str
    composite: int
    model_pick_range: str
    espn_rank: int | None


@dataclass(frozen=True)
class HandbookEntry:
    source_rank: int
    english_name: str
    composite: float
    model_pick_range: str
    model_pick_low: int
    espn_rank: float | None


@dataclass(frozen=True)
class RawTable:
    rows_by_name: dict[str, dict[str, Any]]
    raw_rows: int
    duplicate_names: list[str]


@dataclass
class ProspectBuild:
    pool: PoolProspect
    row: dict[str, Any]
    shooting_pct: float | None
    raw_height_in: float | None
    raw_wingspan_in: float | None
    max_vertical_in: float | None
    hand_length_in: float | None
    unknown_country: bool
    fallback_score: float = -1.0


HANDBOOK_MAPPINGS = [
    HandbookMapping("AJ Dybantsa", 1, "AJ 迪班萨", 52, "1-2", 1),
    HandbookMapping("Darryn Peterson", 51, "达林 彼得森", 44, "1-3", 2),
    HandbookMapping("Cameron Boozer", 15, "卡梅隆 布泽尔", 53, "1-2", 3),
    HandbookMapping("Caleb Wilson", 13, "凯莱布 威尔逊", 48, "3-5", 4),
    HandbookMapping("Keaton Wagler", 19, "基顿 瓦格勒", 45, "4-8", 5),
    HandbookMapping("Darius Acuff", 52, "达柳斯 阿卡夫", 50, "3-6", 6),
    HandbookMapping("K.Flemings", 53, "金斯顿 弗莱明斯", 46, "6-12", 7),
    HandbookMapping("Nate Ament", 84, "纳撒尼尔 阿门特", 46, "7-14", 8),
    HandbookMapping("Y.Lendeborg", 4, "亚克塞尔 伦德伯格", 44, "12-20", 12),
    HandbookMapping("Aday Mara", 55, "阿达伊 马拉", 47, "10-18", 13),
    HandbookMapping("H.Steinbach", 39, "汉内斯 施泰因巴赫", 47, "10-18", 14),
    HandbookMapping("L.Philon", 22, "小拉巴伦 菲隆", 45, "12-20", 15),
    HandbookMapping("Isaiah Evans", 7, "以赛亚 埃文斯", 43, "18-25", 19),
    HandbookMapping("Ebuka Okorie", 17, "埃布卡 奥科里", 41, "25-35", None),
    HandbookMapping("M.Johnson", 23, "小莫雷兹 约翰逊", 47, "15-25", 60),
    HandbookMapping("C.Anderson", 11, "克里斯蒂安 安德森", 42, "20-30", 66),
    HandbookMapping("Cameron Carr", 14, "卡梅隆 卡尔", 43, "20-30", 48),
    HandbookMapping("Meleek Thomas", 56, "马利克 托马斯", 44, "20-30", 26),
    HandbookMapping("Koa Peat", 46, "科阿 皮特", 44, "25-35", None),
    HandbookMapping("Dailyn Swain", 32, "戴林 斯温", 43, "25-40", 51),
    HandbookMapping("J.Quaintance", 38, "杰登 昆坦斯", 29, "35-50", 18),
]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace(ZERO_WIDTH_NON_JOINER, "").strip()


def _is_missing(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and value != value:
        return True
    return clean_text(value).lower() in MISSING_TEXT


def parse_feet_inches(value: Any) -> float | None:
    if _is_missing(value):
        return None
    if isinstance(value, int | float):
        return float(value)
    match = FEET_INCHES_RE.fullmatch(clean_text(value))
    if not match:
        return None
    feet = int(match.group(1))
    inches = float(match.group(2))
    return feet * 12 + inches


def parse_number(value: Any) -> float | None:
    if _is_missing(value):
        return None
    if isinstance(value, int | float):
        return float(value)
    text = clean_text(value).replace(",", "")
    try:
        if text.endswith("%"):
            return float(text[:-1]) / 100
        return float(text)
    except ValueError:
        return None


def ingest_official(
    source_dir: Path,
    out_dir: Path,
    dossier_path: Path | None = None,
    use_llm_divergence: bool = True,
) -> dict[str, Any]:
    source_dir = Path(source_dir)
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    prospects_path = source_dir / "prospects_2627.xlsx"
    answer_card_path = source_dir / "answer_card_template.xlsx"
    handbook_path = source_dir / "handbook.xlsx"

    pool = _read_pool(answer_card_path)
    pool_by_index = {prospect.pool_index: prospect for prospect in pool}
    entrant_table = _read_named_table(prospects_path, "参选人名单", "姓名")
    measurement_table = _read_named_table(prospects_path, "人体测量数据", "球员")
    athletic_table = _read_named_table(prospects_path, "力量与敏捷性", "球员")
    shooting_table = _read_named_table(prospects_path, "投篮训练", "球员")
    raw_tables = {
        "参选人名单": entrant_table,
        "人体测量数据": measurement_table,
        "力量与敏捷性": athletic_table,
        "投篮训练": shooting_table,
    }

    handbook_entries = _read_handbook_entries(handbook_path)
    board_by_pool_index = _map_handbook_to_pool(handbook_entries)
    builds = _build_prospects(
        pool=pool,
        entrants=entrant_table.rows_by_name,
        measurements=measurement_table.rows_by_name,
        athletic=athletic_table.rows_by_name,
        shooting=shooting_table.rows_by_name,
        board_by_pool_index=board_by_pool_index,
    )
    _assign_consensus_ranks(
        builds,
        out_dir=out_dir,
        use_llm_divergence=use_llm_divergence,
    )

    prospect_rows = [build.row for build in sorted(builds, key=lambda item: item.pool.pool_index)]
    draft_order_rows = _read_draft_order(answer_card_path)
    dossiers = _load_dossiers_if_available(dossier_path)
    team_need_rows: list[dict[str, Any]] = []
    mock_signal_rows: list[dict[str, Any]] = []
    if dossiers:
        team_need_rows = _build_team_need_rows(draft_order_rows, dossiers)
        mock_signal_rows = _build_mock_signal_rows(draft_order_rows, prospect_rows)
    q6_options = _read_q6_options(answer_card_path)
    milestone_questions = _read_milestone_questions(answer_card_path)
    divergence_records = _divergence_records(builds)

    _write_csv(out_dir / "prospects.csv", PROSPECT_COLUMNS, prospect_rows)
    _write_csv(out_dir / "draft_order.csv", DRAFT_ORDER_COLUMNS, draft_order_rows)
    _write_csv(out_dir / "team_needs.csv", TEAM_NEEDS_COLUMNS, team_need_rows)
    _write_csv(out_dir / "mock_signals.csv", MOCK_SIGNALS_COLUMNS, mock_signal_rows)
    _write_json(out_dir / "q6_options.json", q6_options)
    _write_json(out_dir / "divergence.json", divergence_records)

    report = _build_report(
        source_dir=source_dir,
        out_dir=out_dir,
        pool=pool,
        pool_by_index=pool_by_index,
        raw_tables=raw_tables,
        measurement_rows=measurement_table.rows_by_name,
        builds=builds,
        handbook_entries=handbook_entries,
        board_by_pool_index=board_by_pool_index,
        q6_options=q6_options,
        milestone_questions=milestone_questions,
        divergence_records=divergence_records,
        dossier_count=len(dossiers),
        team_need_rows=len(team_need_rows),
        mock_signal_rows=len(mock_signal_rows),
    )
    _write_json(out_dir / "ingest_report.json", report)
    return report


def _read_sheet_rows(path: Path, sheet_name: str, header_row: int = 1) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        worksheet = workbook[sheet_name]
        header_values = next(
            worksheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True)
        )
        headers = [clean_text(header) for header in header_values]
        rows: list[dict[str, Any]] = []
        for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
            if all(_is_missing(value) for value in row):
                continue
            rows.append(
                {
                    header: row[index] if index < len(row) else None
                    for index, header in enumerate(headers)
                    if header
                }
            )
        return rows
    finally:
        workbook.close()


def _read_pool(path: Path) -> list[PoolProspect]:
    rows = _read_sheet_rows(path, "球员池_勿修改")
    pool: list[PoolProspect] = []
    for row in rows:
        pool_index_value = parse_number(row.get("序号"))
        name = clean_text(row.get("姓名"))
        if pool_index_value is None or not name:
            continue
        pool_index = int(pool_index_value)
        pool.append(
            PoolProspect(
                pool_index=pool_index,
                prospect_id=f"p{pool_index:03d}",
                name=name,
                position=clean_text(row.get("位置")),
                school=clean_text(row.get("学校/俱乐部")),
                raw_country=clean_text(row.get("国家/地区")),
            )
        )
    return sorted(pool, key=lambda prospect: prospect.pool_index)


def _read_named_table(path: Path, sheet_name: str, key_column: str) -> RawTable:
    rows = _read_sheet_rows(path, sheet_name)
    rows_by_name: dict[str, dict[str, Any]] = {}
    duplicates: list[str] = []
    for row in rows:
        name = clean_text(row.get(key_column))
        if not name:
            continue
        if name in rows_by_name:
            duplicates.append(name)
            continue
        rows_by_name[name] = row
    return RawTable(
        rows_by_name=rows_by_name,
        raw_rows=sum(1 for row in rows if clean_text(row.get(key_column))),
        duplicate_names=sorted(set(duplicates)),
    )


def _read_handbook_entries(path: Path) -> dict[str, HandbookEntry]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        worksheet = workbook["2026届实战应用"]
        entries: dict[str, HandbookEntry] = {}
        for row in worksheet.iter_rows(min_row=4, values_only=True):
            source_rank = parse_number(row[0])
            english_name = clean_text(row[1])
            if source_rank is None or not english_name:
                continue
            composite = parse_number(row[8])
            model_pick_range = clean_text(row[9])
            model_pick_low = _parse_model_pick_low(model_pick_range)
            if composite is None or model_pick_low is None:
                continue
            entries[english_name] = HandbookEntry(
                source_rank=int(source_rank),
                english_name=english_name,
                composite=composite,
                model_pick_range=model_pick_range,
                model_pick_low=model_pick_low,
                espn_rank=parse_number(row[10]),
            )
        return entries
    finally:
        workbook.close()


def _map_handbook_to_pool(
    handbook_entries: dict[str, HandbookEntry],
) -> dict[int, HandbookEntry]:
    mapped: dict[int, HandbookEntry] = {}
    for mapping in HANDBOOK_MAPPINGS:
        entry = handbook_entries.get(mapping.english_name)
        if entry is None:
            entry = HandbookEntry(
                source_rank=0,
                english_name=mapping.english_name,
                composite=float(mapping.composite),
                model_pick_range=mapping.model_pick_range,
                model_pick_low=_parse_model_pick_low(mapping.model_pick_range) or 999,
                espn_rank=None if mapping.espn_rank is None else float(mapping.espn_rank),
            )
        mapped[mapping.pool_index] = entry
    return mapped


def _parse_model_pick_low(value: str) -> int | None:
    match = re.search(r"\d+", value)
    if not match:
        return None
    return int(match.group(0))


def _build_prospects(
    pool: list[PoolProspect],
    entrants: dict[str, dict[str, Any]],
    measurements: dict[str, dict[str, Any]],
    athletic: dict[str, dict[str, Any]],
    shooting: dict[str, dict[str, Any]],
    board_by_pool_index: dict[int, HandbookEntry],
) -> list[ProspectBuild]:
    builds: list[ProspectBuild] = []
    for prospect in pool:
        primary_position = _map_primary_position(prospect.position)
        defaults = POSITION_DEFAULTS[primary_position]
        entrant_row = entrants.get(prospect.name, {})
        measurement_row = measurements.get(prospect.name, {})
        athletic_row = athletic.get(prospect.name, {})
        shooting_row = shooting.get(prospect.name, {})
        board_entry = board_by_pool_index.get(prospect.pool_index)

        age = parse_number(entrant_row.get("年龄"))
        barefoot_height_in = parse_feet_inches(measurement_row.get("无鞋身高"))
        wingspan_in = parse_feet_inches(measurement_row.get("臂展"))
        standing_reach_in = parse_feet_inches(measurement_row.get("摸高"))
        hand_length_in = parse_number(measurement_row.get("手长(英寸)"))
        hand_width_in = parse_number(measurement_row.get("手宽(英寸)"))
        weight_lb = parse_number(measurement_row.get("体重(磅)"))
        max_vertical_in = parse_number(athletic_row.get("助跑弹跳(英寸)"))
        standing_vertical_in = parse_number(athletic_row.get("原地弹跳(英寸)"))
        shooting_pct = _shooting_pct(shooting_row)
        country, is_international, unknown_country = _resolve_country(prospect)

        row = {
            "prospect_id": prospect.prospect_id,
            "name": prospect.name,
            "primary_position": primary_position,
            "archetype": ARCHETYPES[primary_position],
            "consensus_rank": None,
            "age": age if age is not None else 20.0,
            "height_in": (
                barefoot_height_in if barefoot_height_in is not None else defaults["height_in"]
            ),
            "wingspan_in": wingspan_in if wingspan_in is not None else defaults["wingspan_in"],
            "usage_rate": defaults["usage_rate"],
            "true_shooting_pct": (
                shooting_pct if shooting_pct is not None else defaults["true_shooting_pct"]
            ),
            "assist_rate": defaults["assist_rate"],
            "rebound_rate": defaults["rebound_rate"],
            "stock_rate": defaults["stock_rate"],
            "notes": "official-2026",
            "barefoot_height_in": barefoot_height_in,
            "hand_length_in": hand_length_in,
            "hand_width_in": hand_width_in,
            "standing_reach_in": standing_reach_in,
            "weight_lb": weight_lb,
            "max_vertical_in": max_vertical_in,
            "standing_vertical_in": standing_vertical_in,
            "school": "" if _is_missing(prospect.school) else prospect.school,
            "country": country,
            "is_international": is_international,
            "is_center": prospect.position in {"中锋", "C"},
            "talent_composite": None if board_entry is None else board_entry.composite,
            "espn_rank": (
                None
                if board_entry is None or board_entry.espn_rank is None
                else int(board_entry.espn_rank)
            ),
            "model_pick_low": None if board_entry is None else board_entry.model_pick_low,
            "board_source": "handbook" if board_entry is not None else "fallback",
            "talent_rank": None,
            "market_rank": None,
            "talent_signal": None,
            "market_signal": None,
            "divergence_gap": None,
            "divergence_type": "",
            "divergence_reason": "",
            "divergence_llm_verdict": "",
            "divergence_llm_market_weight": None,
            "divergence_llm_confidence": None,
            "divergence_llm_reasoning": "",
            "fused_score": None,
        }
        builds.append(
            ProspectBuild(
                pool=prospect,
                row=row,
                shooting_pct=shooting_pct,
                raw_height_in=barefoot_height_in,
                raw_wingspan_in=wingspan_in,
                max_vertical_in=max_vertical_in,
                hand_length_in=hand_length_in,
                unknown_country=unknown_country,
            )
        )
    return builds


def _map_primary_position(position: str) -> str:
    try:
        return POSITION_TO_PRIMARY[position]
    except KeyError as exc:
        raise ValueError(f"Unknown prospect position: {position}") from exc


def _resolve_country(prospect: PoolProspect) -> tuple[str, bool, bool]:
    country = clean_text(prospect.raw_country)
    if _is_missing(country):
        school = clean_text(prospect.school)
        if school in COUNTRY_NAMES:
            country = school
        else:
            return "", False, True
    return country, country != "美国", False


def _shooting_pct(row: dict[str, Any]) -> float | None:
    made_total = 0.0
    attempts_total = 0.0
    for made_key, attempt_key in SHOOTING_PAIRS:
        made = parse_number(row.get(made_key))
        attempts = parse_number(row.get(attempt_key))
        if made is None or attempts is None or attempts <= 0:
            continue
        made_total += made
        attempts_total += attempts
    if attempts_total <= 0:
        return None
    return made_total / attempts_total


def _assign_consensus_ranks(
    builds: list[ProspectBuild],
    *,
    out_dir: Path | None = None,
    use_llm_divergence: bool = True,
) -> None:
    _assign_fallback_scores(builds)
    handbook_builds = [build for build in builds if build.row["board_source"] == "handbook"]
    fallback_builds = [build for build in builds if build.row["board_source"] == "fallback"]

    _assign_talent_ranks(builds, handbook_builds, fallback_builds)
    _assign_market_signals_and_divergence(builds)
    _apply_llm_divergence(builds, out_dir, use_llm_divergence)
    llm_divergence_active = use_llm_divergence and out_dir is not None

    for build in handbook_builds:
        talent_signal = float(build.row["talent_signal"])
        market_signal = build.row["market_signal"]
        llm_market_weight = build.row.get("divergence_llm_market_weight")
        if market_signal is None:
            fused_score = talent_signal
        elif llm_divergence_active and llm_market_weight is not None:
            confidence = float(build.row.get("divergence_llm_confidence") or 0.0)
            w_rule = 0.6 if build.row["divergence_type"] == "aligned" else 0.5
            w = w_rule + confidence * (float(llm_market_weight) - w_rule)
            w = min(1.0, max(0.0, w))
            fused_score = w * float(market_signal) + (1 - w) * talent_signal
            build.row["divergence_reason"] = _append_llm_divergence_reason(build.row)
        elif build.row["divergence_type"] == "aligned":
            fused_score = 0.6 * float(market_signal) + 0.4 * talent_signal
        else:
            fused_score = 0.5 * float(market_signal) + 0.5 * talent_signal
        build.row["fused_score"] = fused_score

    for build in fallback_builds:
        build.row["fused_score"] = build.fallback_score

    handbook_builds.sort(
        key=lambda build: (-float(build.row["fused_score"]), build.pool.prospect_id)
    )
    fallback_builds.sort(key=lambda build: (-build.fallback_score, build.pool.pool_index))

    for rank, build in enumerate([*handbook_builds, *fallback_builds], start=1):
        build.row["consensus_rank"] = rank


def _assign_talent_ranks(
    builds: list[ProspectBuild],
    handbook_builds: list[ProspectBuild],
    fallback_builds: list[ProspectBuild],
) -> None:
    denominator = max(len(builds) - 1, 1)
    handbook_ranked = sorted(
        handbook_builds,
        key=lambda build: (-float(build.row["talent_composite"]), build.pool.prospect_id),
    )
    fallback_ranked = sorted(
        fallback_builds,
        key=lambda build: (-build.fallback_score, build.pool.prospect_id),
    )
    for rank, build in enumerate(handbook_ranked, start=1):
        build.row["talent_rank"] = rank
        build.row["talent_signal"] = 1 - ((rank - 1) / denominator)
    for rank, build in enumerate(fallback_ranked, start=len(handbook_ranked) + 1):
        build.row["talent_rank"] = rank
        build.row["talent_signal"] = 1 - ((rank - 1) / denominator)


def _assign_market_signals_and_divergence(builds: list[ProspectBuild]) -> None:
    espn_ranks = [
        int(build.row["espn_rank"])
        for build in builds
        if build.row.get("espn_rank") is not None
    ]
    espn_max = max(espn_ranks) if espn_ranks else 1
    denominator = max(espn_max - 1, 1)

    for build in builds:
        espn_rank = build.row.get("espn_rank")
        if espn_rank is not None:
            market_rank = int(espn_rank)
            build.row["market_rank"] = market_rank
            build.row["market_signal"] = 1 - ((market_rank - 1) / denominator)
        if build.row["board_source"] != "handbook" or espn_rank is None:
            continue

        talent_rank = int(build.row["talent_rank"])
        market_rank = int(espn_rank)
        gap = market_rank - talent_rank
        build.row["divergence_gap"] = gap
        build.row["divergence_type"] = _divergence_type(gap)
        build.row["divergence_reason"] = _divergence_reason(
            name=build.pool.name,
            talent_rank=talent_rank,
            espn_rank=market_rank,
            gap=gap,
            divergence_type=str(build.row["divergence_type"]),
        )


def _neutral_divergence_notes(build: ProspectBuild) -> str:
    """Fact-only context for the LLM adjudicator.

    Deliberately omits the deterministic verdict (market_hype/market_fade) so the
    model resolves the split on the merits instead of being anchored to the rule
    label it is meant to second-guess.
    """
    row = build.row
    talent_rank = int(row["talent_rank"])
    market_rank = float(row["market_rank"])
    gap = int(row["divergence_gap"])
    direction = (
        "the public market is markedly more bullish than the structured talent model"
        if gap < 0
        else "the structured talent model is markedly more bullish than the public market"
    )
    return (
        f"{build.pool.name}: talent-model rank #{talent_rank} vs market rank "
        f"#{market_rank:g} (gap {gap:+d}); {direction}. Decide which signal to trust "
        "from the measurables and production, without presuming either is correct."
    )


def _apply_llm_divergence(
    builds: list[ProspectBuild],
    out_dir: Path | None,
    use_llm_divergence: bool,
) -> None:
    if not use_llm_divergence or out_dir is None:
        return

    cache_path = Path(out_dir) / "divergence_llm.json"
    cache = _load_llm_divergence_cache(cache_path)

    for build in builds:
        if not _needs_llm_divergence(build, cache):
            continue
        result = divergence.reason_divergence(
            name=build.pool.name,
            position=str(build.row["primary_position"]),
            talent_profile={
                "talent_composite": build.row.get("talent_composite"),
                "true_shooting_pct": build.row.get("true_shooting_pct"),
                "usage_rate": build.row.get("usage_rate"),
                "assist_rate": build.row.get("assist_rate"),
                "rebound_rate": build.row.get("rebound_rate"),
                "stock_rate": build.row.get("stock_rate"),
                "age": build.row.get("age"),
                "height_in": build.row.get("height_in"),
                "wingspan_in": build.row.get("wingspan_in"),
                "archetype": build.row.get("archetype"),
            },
            talent_rank=int(build.row["talent_rank"]),
            market_rank=float(build.row["market_rank"]),
            divergence=int(build.row["divergence_gap"]),
            notes=_neutral_divergence_notes(build),
        )
        entry = _coerce_llm_divergence_entry(result)
        if entry is not None:
            cache[build.pool.prospect_id] = entry

    _write_llm_divergence_cache(cache_path, cache)

    for build in builds:
        entry = cache.get(build.pool.prospect_id)
        if entry is None:
            build.row["divergence_llm_verdict"] = ""
            build.row["divergence_llm_market_weight"] = None
            build.row["divergence_llm_confidence"] = None
            build.row["divergence_llm_reasoning"] = ""
            continue
        build.row["divergence_llm_verdict"] = entry["verdict"]
        build.row["divergence_llm_market_weight"] = entry["adjusted_market_weight"]
        build.row["divergence_llm_confidence"] = entry["confidence"]
        build.row["divergence_llm_reasoning"] = entry["reasoning"]


def _needs_llm_divergence(
    build: ProspectBuild,
    cache: dict[str, dict[str, Any]],
) -> bool:
    if build.pool.prospect_id in cache:
        return False
    if build.row.get("board_source") != "handbook":
        return False
    if build.row.get("espn_rank") is None:
        return False
    gap = build.row.get("divergence_gap")
    return gap is not None and abs(int(gap)) >= 8


def _load_llm_divergence_cache(path: Path) -> dict[str, dict[str, Any]]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    if not isinstance(payload, dict):
        return {}

    cache: dict[str, dict[str, Any]] = {}
    for prospect_id, entry in payload.items():
        if not isinstance(prospect_id, str):
            continue
        coerced = _coerce_llm_divergence_entry(entry)
        if coerced is not None:
            cache[prospect_id] = coerced
    return cache


def _write_llm_divergence_cache(
    path: Path,
    cache: dict[str, dict[str, Any]],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(cache, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def _coerce_llm_divergence_entry(entry: Any) -> dict[str, Any] | None:
    if not isinstance(entry, dict):
        return None

    verdict = entry.get("verdict")
    if verdict not in divergence.VERDICTS:
        return None

    market_weight = _unit_interval(entry.get("adjusted_market_weight"))
    confidence = _unit_interval(entry.get("confidence"))
    reasoning = entry.get("reasoning")
    if market_weight is None or confidence is None or not isinstance(reasoning, str):
        return None
    reasoning = reasoning.strip()
    if not reasoning:
        return None

    return {
        "verdict": verdict,
        "adjusted_market_weight": market_weight,
        "confidence": confidence,
        "reasoning": reasoning,
    }


def _unit_interval(value: Any) -> float | None:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if number < 0.0 or number > 1.0:
        return None
    return number


def _append_llm_divergence_reason(row: dict[str, Any]) -> str:
    reason = str(row.get("divergence_reason") or "")
    if "[gpt-5.5:" in reason:
        return reason
    verdict = str(row.get("divergence_llm_verdict") or "")
    market_weight = float(row.get("divergence_llm_market_weight") or 0.0)
    confidence = float(row.get("divergence_llm_confidence") or 0.0)
    reasoning = str(row.get("divergence_llm_reasoning") or "").strip()
    return (
        f"{reason} [gpt-5.5:{verdict} w={market_weight:g} "
        f"conf={confidence:g} — {reasoning}]"
    )


def _divergence_type(gap: int) -> str:
    if abs(gap) <= 10:
        return "aligned"
    if gap > 10:
        return "market_fade"
    return "market_hype"


def _divergence_reason(
    name: str,
    talent_rank: int,
    espn_rank: int,
    gap: int,
    divergence_type: str,
) -> str:
    gap_text = f"+{gap}" if gap > 0 else str(gap)
    if divergence_type == "market_fade":
        return (
            f"{name}:天赋第{talent_rank}但ESPN第{espn_rank}(gap{gap_text}):"
            "数据亮眼、市场退烧的stats-darling,顺位不确定性高。"
        )
    if divergence_type == "market_hype":
        return (
            f"{name}:ESPN第{espn_rank}高于天赋第{talent_rank}(gap{gap_text}):"
            "市场热捧但天赋模型保守,需防顺位过热。"
        )
    return (
        f"{name}:天赋第{talent_rank}与ESPN第{espn_rank}(gap{gap_text}):"
        "双信号基本一致,顺位锚点稳定。"
    )


def _assign_fallback_scores(builds: list[ProspectBuild]) -> None:
    feature_names = [
        "shooting_pct",
        "wingspan_delta",
        "max_vertical_in",
        "raw_height_in",
    ]
    feature_values = {
        "shooting_pct": [build.shooting_pct for build in builds],
        "wingspan_delta": [
            None
            if build.raw_height_in is None or build.raw_wingspan_in is None
            else build.raw_wingspan_in - build.raw_height_in
            for build in builds
        ],
        "max_vertical_in": [build.max_vertical_in for build in builds],
        "raw_height_in": [build.raw_height_in for build in builds],
    }
    normalized = {
        name: _normalize_feature(values)
        for name, values in feature_values.items()
    }
    weights = {
        "shooting_pct": 0.35,
        "wingspan_delta": 0.25,
        "max_vertical_in": 0.20,
        "raw_height_in": 0.20,
    }
    for index, build in enumerate(builds):
        weighted = 0.0
        weight_total = 0.0
        for feature_name in feature_names:
            value = normalized[feature_name][index]
            if value is None:
                continue
            weighted += weights[feature_name] * value
            weight_total += weights[feature_name]
        build.fallback_score = -1.0 if weight_total == 0 else weighted / weight_total


def _normalize_feature(values: list[float | None]) -> list[float | None]:
    present = [value for value in values if value is not None]
    if not present:
        return [None for _ in values]
    low = min(present)
    high = max(present)
    if high == low:
        return [None if value is None else 0.5 for value in values]
    return [None if value is None else (value - low) / (high - low) for value in values]


def _read_draft_order(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        worksheet = workbook["②选秀名单预测"]
        rows: list[dict[str, Any]] = []
        for row in worksheet.iter_rows(min_row=3, values_only=True):
            pick = parse_number(row[0])
            team = clean_text(row[1])
            if pick is None or not team:
                continue
            if team not in TEAM_ABBREVIATIONS:
                raise ValueError(f"Unknown NBA team in draft order: {team}")
            rows.append(
                {
                    "pick": int(pick),
                    "team": team,
                    "abbreviation": TEAM_ABBREVIATIONS[team],
                }
            )
        return rows
    finally:
        workbook.close()


def _load_dossiers_if_available(
    dossier_path: Path | None,
) -> dict[str, TeamDossier]:
    resolved_path = DEFAULT_DOSSIER_PATH if dossier_path is None else Path(dossier_path)
    if not resolved_path.exists():
        return {}
    return load_team_dossiers(resolved_path)


def _build_team_need_rows(
    draft_order_rows: list[dict[str, Any]],
    dossiers: dict[str, TeamDossier],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    seen: set[str] = set()
    for draft_row in draft_order_rows:
        abbreviation = str(draft_row["abbreviation"])
        if abbreviation in seen:
            continue
        seen.add(abbreviation)
        dossier = dossiers.get(abbreviation)
        if dossier is None:
            raise ValueError(f"Missing team dossier for draft-order team: {abbreviation}")
        for position in ("G", "W", "B"):
            rows.append(
                {
                    "abbreviation": abbreviation,
                    "position": position,
                    "weight": dossier.roster_needs[position],
                }
            )
    return rows


def _build_mock_signal_rows(
    draft_order_rows: list[dict[str, Any]],
    prospect_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    team_order: dict[str, int] = {}
    for index, row in enumerate(draft_order_rows):
        abbreviation = str(row["abbreviation"])
        if abbreviation not in team_order:
            team_order[abbreviation] = index
    signal_by_key: dict[tuple[str, str], dict[str, Any]] = {}
    for prospect in prospect_rows:
        prospect_id = str(prospect["prospect_id"])
        anchors = _market_anchors(prospect)
        if not anchors:
            continue
        for draft_row in draft_order_rows:
            pick = int(draft_row["pick"])
            abbreviation = str(draft_row["abbreviation"])
            signal_strength = 0.0
            for _, anchor, base_strength in anchors:
                distance = abs(pick - anchor)
                if distance <= 4:
                    signal_strength = max(signal_strength, base_strength - 0.11 * distance)
            if signal_strength <= 0:
                continue
            key = (abbreviation, prospect_id)
            current = signal_by_key.get(key)
            if current is None or float(current["signal_strength"]) < signal_strength:
                signal_by_key[key] = {
                    "abbreviation": abbreviation,
                    "prospect_id": prospect_id,
                    "signal_strength": _clamp_signal(signal_strength),
                    "source": "handbook-market",
                }
    return sorted(
        signal_by_key.values(),
        key=lambda row: (team_order.get(str(row["abbreviation"]), 999), str(row["prospect_id"])),
    )


def _market_anchors(prospect: dict[str, Any]) -> list[tuple[str, int, float]]:
    anchors: list[tuple[str, int, float]] = []
    espn_rank = prospect.get("espn_rank")
    if espn_rank is not None:
        rank = int(espn_rank)
        if rank <= 35:
            anchors.append(("espn_rank", rank, 0.88))
    model_pick_low = prospect.get("model_pick_low")
    if model_pick_low is not None:
        rank = int(model_pick_low)
        if rank <= 35:
            anchors.append(("model_pick_low", rank, 0.78))
    return anchors


def _clamp_signal(value: float) -> float:
    return max(0.0, min(1.0, value))


def _read_q6_options(path: Path) -> list[str]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        worksheet = workbook["选项数据源"]
        options: list[str] = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            option = clean_text(row[0])
            if option:
                options.append(option)
        return options
    finally:
        workbook.close()


def _read_milestone_questions(path: Path) -> list[dict[str, str]]:
    rows = _read_sheet_rows(path, "③里程碑预测", header_row=2)
    questions: list[dict[str, str]] = []
    for row in rows:
        qid = clean_text(row.get("题号"))
        if not qid:
            continue
        questions.append(
            {
                "id": qid,
                "topic": clean_text(row.get("主题")),
                "question": clean_text(row.get("题目")),
                "answer_format": clean_text(row.get("答案格式")),
                "reveal_timing": clean_text(row.get("揭晓时刻")),
            }
        )
    return questions


def _build_report(
    source_dir: Path,
    out_dir: Path,
    pool: list[PoolProspect],
    pool_by_index: dict[int, PoolProspect],
    raw_tables: dict[str, RawTable],
    measurement_rows: dict[str, dict[str, Any]],
    builds: list[ProspectBuild],
    handbook_entries: dict[str, HandbookEntry],
    board_by_pool_index: dict[int, HandbookEntry],
    q6_options: list[str],
    milestone_questions: list[dict[str, str]],
    divergence_records: list[dict[str, Any]],
    dossier_count: int,
    team_need_rows: int,
    mock_signal_rows: int,
) -> dict[str, Any]:
    builds_by_pool_index = {build.pool.pool_index: build for build in builds}
    table_matches = {
        sheet_name: _table_match_report(pool, raw_table)
        for sheet_name, raw_table in raw_tables.items()
    }
    handbook_mapping_rows = [
        _handbook_mapping_report_row(
            mapping=mapping,
            pool=pool_by_index.get(mapping.pool_index),
            build=builds_by_pool_index.get(mapping.pool_index),
            workbook_entry=handbook_entries.get(mapping.english_name),
            mapped_entry=board_by_pool_index.get(mapping.pool_index),
        )
        for mapping in HANDBOOK_MAPPINGS
    ]
    return {
        "source_dir": str(source_dir),
        "output_dir": str(out_dir),
        "pool_count": len(pool),
        "table_matches": table_matches,
        "handbook_mapping_count": len(handbook_mapping_rows),
        "handbook_mappings": handbook_mapping_rows,
        "parse_samples": _parse_samples(measurement_rows),
        "missing_stats": _missing_stats(builds),
        "market_coverage": _market_coverage(builds),
        "divergence_stats": _divergence_stats(builds),
        "llm_divergence_verdict_count": _llm_divergence_verdict_count(builds),
        "divergence_top5": divergence_records[:5],
        "dossier_count": dossier_count,
        "team_need_rows": team_need_rows,
        "mock_signal_rows": mock_signal_rows,
        "milestone_raw_preview": _milestone_preview(builds),
        "q6_option_count": len(q6_options),
        "milestone_questions": milestone_questions,
        "output_files": [
            "prospects.csv",
            "draft_order.csv",
            "team_needs.csv",
            "mock_signals.csv",
            "q6_options.json",
            "divergence.json",
            "ingest_report.json",
        ],
    }


def _llm_divergence_verdict_count(builds: list[ProspectBuild]) -> int:
    return sum(1 for build in builds if build.row.get("divergence_llm_verdict"))


def _table_match_report(pool: list[PoolProspect], table: RawTable) -> dict[str, Any]:
    missing = [
        {"pool_index": prospect.pool_index, "name": prospect.name}
        for prospect in pool
        if prospect.name not in table.rows_by_name
    ]
    return {
        "raw_rows": table.raw_rows,
        "raw_unique_names": len(table.rows_by_name),
        "matched_pool_rows": len(pool) - len(missing),
        "missing_pool_rows": len(missing),
        "unmatched_pool_names": missing,
        "duplicate_raw_names": table.duplicate_names,
    }


def _handbook_mapping_report_row(
    mapping: HandbookMapping,
    pool: PoolProspect | None,
    build: ProspectBuild | None,
    workbook_entry: HandbookEntry | None,
    mapped_entry: HandbookEntry | None,
) -> dict[str, Any]:
    flags: list[str] = []
    notes: list[str] = []
    if pool is None:
        flags.append("pool_index_missing")
    elif pool.name != mapping.chinese_name:
        flags.append("pool_name_mismatch")
    if workbook_entry is None:
        flags.append("handbook_row_missing")
    else:
        if int(workbook_entry.composite) != mapping.composite:
            flags.append("composite_mismatch")
        if workbook_entry.model_pick_range != mapping.model_pick_range:
            flags.append("model_range_mismatch")
        workbook_espn = None if workbook_entry.espn_rank is None else int(workbook_entry.espn_rank)
        if workbook_espn != mapping.espn_rank:
            flags.append("espn_rank_mismatch")
    if mapping.english_name == "Nate Ament":
        notes.append("duplicate_chinese_name_forced_pool_84")
    return {
        "english_name": mapping.english_name,
        "pool_index": mapping.pool_index,
        "mapped_chinese_name": mapping.chinese_name,
        "actual_pool_name": None if pool is None else pool.name,
        "actual_school": None if pool is None else pool.school,
        "actual_position": None if pool is None else pool.position,
        "talent_composite": None if mapped_entry is None else mapped_entry.composite,
        "model_pick_range": None if mapped_entry is None else mapped_entry.model_pick_range,
        "model_pick_low": None if mapped_entry is None else mapped_entry.model_pick_low,
        "espn_rank": None if mapped_entry is None else mapped_entry.espn_rank,
        "consensus_rank": None if build is None else build.row["consensus_rank"],
        "status": "review" if flags else "ok",
        "flags": flags,
        "notes": notes,
    }


def _parse_samples(measurement_rows: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    wanted_raw_values = ["6' 3.75''", "7' 3.25''", "6' 9.50''"]
    samples: list[dict[str, Any]] = []
    for wanted in wanted_raw_values:
        for player, row in measurement_rows.items():
            for field in ("无鞋身高", "臂展"):
                raw_value = clean_text(row.get(field))
                if raw_value == wanted:
                    samples.append(
                        {
                            "player": player,
                            "field": field,
                            "raw": raw_value,
                            "inches": parse_feet_inches(raw_value),
                        }
                    )
                    break
            if len(samples) == len(wanted_raw_values) or (
                samples and samples[-1]["raw"] == wanted
            ):
                break
    if len(samples) >= 3:
        return samples[:3]

    seen = {sample["raw"] for sample in samples}
    for player, row in measurement_rows.items():
        for field in ("无鞋身高", "臂展"):
            raw_value = clean_text(row.get(field))
            if raw_value and raw_value not in seen:
                samples.append(
                    {
                        "player": player,
                        "field": field,
                        "raw": raw_value,
                        "inches": parse_feet_inches(raw_value),
                    }
                )
                seen.add(raw_value)
            if len(samples) == 3:
                return samples
    return samples


def _market_coverage(builds: list[ProspectBuild]) -> dict[str, Any]:
    market_count = sum(1 for build in builds if build.row.get("market_rank") is not None)
    handbook_builds = [build for build in builds if build.row["board_source"] == "handbook"]
    handbook_market_count = sum(
        1 for build in handbook_builds if build.row.get("market_rank") is not None
    )
    return {
        "pool_count": len(builds),
        "market_rank_count": market_count,
        "market_rank_coverage": market_count / len(builds) if builds else 0.0,
        "handbook_count": len(handbook_builds),
        "handbook_market_rank_count": handbook_market_count,
        "handbook_market_rank_coverage": (
            handbook_market_count / len(handbook_builds) if handbook_builds else 0.0
        ),
    }


def _divergence_stats(builds: list[ProspectBuild]) -> dict[str, int]:
    types = ["aligned", "market_fade", "market_hype"]
    return {
        divergence_type: sum(
            1
            for build in builds
            if build.row.get("divergence_type") == divergence_type
        )
        for divergence_type in types
    }


def _divergence_records(builds: list[ProspectBuild]) -> list[dict[str, Any]]:
    records = []
    for build in builds:
        divergence_type = build.row.get("divergence_type")
        if divergence_type not in {"market_fade", "market_hype"}:
            continue
        records.append(
            {
                "prospect_id": build.pool.prospect_id,
                "name": build.pool.name,
                "school": build.row["school"],
                "board_source": build.row["board_source"],
                "consensus_rank": build.row["consensus_rank"],
                "talent_rank": build.row["talent_rank"],
                "market_rank": build.row["market_rank"],
                "talent_signal": build.row["talent_signal"],
                "market_signal": build.row["market_signal"],
                "fused_score": build.row["fused_score"],
                "divergence_gap": build.row["divergence_gap"],
                "divergence_type": divergence_type,
                "divergence_reason": build.row["divergence_reason"],
                "divergence_llm_verdict": build.row.get("divergence_llm_verdict", ""),
                "divergence_llm_market_weight": build.row.get(
                    "divergence_llm_market_weight"
                ),
                "divergence_llm_confidence": build.row.get("divergence_llm_confidence"),
                "divergence_llm_reasoning": build.row.get("divergence_llm_reasoning", ""),
            }
        )
    return sorted(
        records,
        key=lambda record: (
            -abs(int(record["divergence_gap"])),
            str(record["divergence_type"]),
            str(record["prospect_id"]),
        ),
    )


def _missing_stats(builds: list[ProspectBuild]) -> dict[str, int]:
    return {
        "height_and_wingspan_count": sum(
            1
            for build in builds
            if build.raw_height_in is not None and build.raw_wingspan_in is not None
        ),
        "max_vertical_count": sum(1 for build in builds if build.max_vertical_in is not None),
        "hand_length_count": sum(1 for build in builds if build.hand_length_in is not None),
    }


def _milestone_preview(builds: list[ProspectBuild]) -> dict[str, Any]:
    wingspan_delta_count = sum(
        1
        for build in builds
        if build.raw_height_in is not None
        and build.raw_wingspan_in is not None
        and build.raw_wingspan_in - build.raw_height_in >= 5.0
    )
    return {
        "wingspan_minus_barefoot_height_gte_5_count": wingspan_delta_count,
        "center_count": sum(1 for build in builds if build.row["is_center"]),
        "international_count": sum(1 for build in builds if build.row["is_international"]),
        "unknown_country_count": sum(1 for build in builds if build.unknown_country),
        "top3_max_vertical": _top_metric(builds, "max_vertical_in", 3),
        "top5_hand_length": _top_metric(builds, "hand_length_in", 5),
    }


def _top_metric(
    builds: list[ProspectBuild],
    row_key: str,
    limit: int,
) -> list[dict[str, Any]]:
    ranked = [
        build
        for build in builds
        if build.row.get(row_key) is not None
    ]
    ranked.sort(key=lambda build: (-float(build.row[row_key]), build.pool.pool_index))
    return [
        {
            "pool_index": build.pool.pool_index,
            "prospect_id": build.pool.prospect_id,
            "name": build.pool.name,
            "value": build.row[row_key],
        }
        for build in ranked[:limit]
    ]


def _write_csv(path: Path, columns: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow({column: _format_csv_value(row.get(column)) for column in columns})


def _write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )


def _format_csv_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float):
        if value != value:
            return ""
        return f"{value:.4f}".rstrip("0").rstrip(".")
    return str(value)
